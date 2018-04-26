import imaplib
import email
import email.message
import time
import mimetypes
import base64
import email.header
import quopri
import zipfile
import json
import os
import sys
import csv
from dbfread import DBF
import glob
from xlwt import Workbook
import xlwt, xlrd
from xlutils.copy import copy as xlcopy
from xlwt import Workbook
import subprocess
import re
import shutil
import fnmatch
import ftplib
import openpyxl as xl
from xml.sax import ContentHandler, parse
import pandas as pd

Pharmacy = 'pharmacy.json'
Mail_login = 'm_login.json'
FTP_login = 'f_login.json'

text = open('zvit.txt', 'a')

class ExcelHandler(ContentHandler):
    def __init__(self):
        self.chars = [  ]
        self.cells = [  ]
        self.rows = [  ]
        self.tables = [  ]
    def characters(self, content):
        self.chars.append(content)
    def startElement(self, name, atts):
        if name=="Cell":
            self.chars = [  ]
        elif name=="Row":
            self.cells=[  ]
        elif name=="Table":
            self.rows = [  ]
    def endElement(self, name):
        if name=="Cell":
            self.cells.append(''.join(self.chars))
        elif name=="Row":
            self.rows.append(self.cells)
        elif name=="Table":
            self.tables.append(self.rows)

def json_f(ph):
    with open(ph, encoding='utf-8') as data_file:
        return json.loads(data_file.read())


def ftp_l(id, file_ftp):
    ftp_login = json_f(FTP_login)
    FTP_HOST = ftp_login['FTP'][0]['FTP_HOST']
    FTP_USER = ftp_login['FTP'][0]['FTP_USER'    ]
    FTP_PASSWORD = ftp_login['FTP'][0]['FTP_PASSWORD']
    ftp = ftplib.FTP(FTP_HOST, FTP_USER, FTP_PASSWORD)
    fin = open('%s' % file_ftp, 'rb')
    ftp.cwd('/%s' % id)
    ftp.storbinary('STOR %s' % id  + '_d.xls', fin)
    text.write(time.strftime("%b_%d_%Y_%H_%M_") + '  ' + 'ftp_load' + file_ftp + '\n' +'\n')
    ftp.quit()


def code_en(s_filename):
    encodings = ['utf-8', 'windows-1251', 'windows-1252', 'utf-16']
    for e in encodings:
        try:
            read_book = xlrd.open_workbook(s_filename, encoding_override=e, on_demand=True)
            read_book.sheet_by_index(0)
        except UnicodeDecodeError:
            print('got unicode error with %s , trying different encoding' % e)
        else:
            return e


def xls_rows(id, js_xl, source_filename):
    destination_filename = source_filename[:source_filename.rfind('.')] + '_d' + '.xls'
    read_book = xlrd.open_workbook(source_filename, encoding_override=code_en(source_filename), on_demand=True)
    read_sheet = read_book.sheet_by_index(0)
    write_book = Workbook()
    sheet1 = write_book.add_sheet('Sheet 1')
    for row_index in range(read_sheet.nrows):
        a = row_index + int(js_xl[0]['first_c'])
        b = row_index + 1
        if a == read_sheet.nrows:
            break
        name_r = js_xl[0]['name_r']
        amount_r = js_xl[0]['amount_r']
        sheet1.write(b, 0, read_sheet.cell(a, int(name_r)).value)
        sheet1.write(b, 1, read_sheet.cell(a, int(amount_r)).value)

    write_book.save(destination_filename)
    print(destination_filename)
    ftp_l(id, destination_filename)


def xml_xls(id, js_xm, xm_file):
    excelHandler = ExcelHandler()
    parse(xm_file, excelHandler)
    df1 = pd.DataFrame(excelHandler.tables[0][4:], columns=excelHandler.tables[0][0])
    new_name = xm_file[:-4] + '_x.xls'
    writer = pd.ExcelWriter(new_name)
    df1.to_excel(writer, sheet_name='Sheet1')
    writer.save()
    xls_rows(id, js_xm, new_name)


def csv_xls(id, js_cs, csv_file):
    workbook = Workbook()
    worksheet = workbook.add_sheet('Sheet 1')
    xls_name = csv_file[:-4] + '.xls'
    with open(csv_file, 'rt', encoding='utf8') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                worksheet.write(r, c, col)
    workbook.save(xls_name)
    xls_rows(id, js_cs, xls_name)


def dbf_csv(id, js_db, name):
    table = DBF(name)
    csv_fn = name[:-4] + ".csv"

    with open(csv_fn, 'w') as f:
        writer = csv.writer(f)
        writer.writerow(table.field_names)
        for record in table:
            writer.writerow(list(record.values()))
    csv_xls(id, js_db, csv_fn)


def unzip_file(id, js_uf, f_name):
    with zipfile.ZipFile(f_name, 'r') as zip_f:
        for name in zip_f.namelist():
            new_name = f_name[:-4] + name[name.rfind('.'):]

            with zip_f.open(name) as file:
                content = file.read()
                full_path = os.path.join(os.getcwd(), new_name)
                with open(full_path, 'wb') as file_w:
                    file_w.write(content)
                print(new_name)
                type_file(id, js_uf, new_name)


def type_file(id, js_tf, filename):
    if filename[filename.rfind('.'):].lower() == '.zip':
        unzip_file(id, js_tf, filename)
        print(filename + '_uzip')
    elif filename[filename.rfind('.'):].lower() == '.dbf':
        dbf_csv(id, js_tf, filename)
        print(filename + '_conv_dbf')
    elif filename[filename.rfind('.'):].lower() == '.csv':
        csv_xls(id, js_tf, filename)
        print(filename + '_conv_csv')
    elif filename[filename.rfind('.'):].lower() == '.xls' or '.xlsx':
        try:
            xlrd.open_workbook(filename)
        except Exception as er:
            if str(er).rfind('xml') > 0:
                xml_xls(id, js_tf, filename)
                print(filename + '_read_xml')
        else:
            xls_rows(id, js_tf, filename)
            print(filename + '_read_xls')
    else:
        print('Bad file' + filename)


def re_name(js_rn, file_name):

    for i in js_rn[0]['id']:
        print(i)
        if fnmatch.fnmatch(file_name, '*%s*' % i):
            id_n = js_rn[0]['id'][i]
            file_p = './mail/' + js_rn[0]['client'] + '/working/'
            file_n = shutil.copy2(file_name, file_p + id_n + file_name[file_name.rfind('.'):])
            type_file(id_n, js_rn, file_n)
            print(file_n)


def main_loop(js_d):
    login = json_f(Mail_login)[js_d[0]['mail']]
    print(json_f(Mail_login))
    print(login)
    Miro_HOST = login[0]['Miro_HOST']
    print(Miro_HOST)
    Miro_PORT = login[0]['Miro_PORT']
    Miro_USER = login[0]['Miro_USER']
    Miro_PASSWORD = login[0]['Miro_PASSWORD']
    print("Connecting to {}...".format(Miro_HOST))
    imap4 = imaplib.IMAP4_SSL(host=Miro_HOST, port=Miro_PORT)
    print("Connected! Logging in as {}...".format(Miro_USER))
    imap4.login(user=Miro_USER, password=Miro_PASSWORD)
    print("Logged in! Listing messages...")
    imap4.select('INBOX')
    # nmessages = select_data[0].decode('utf-8')
    status, search_data = imap4.search(None, 'FROM', js_d[0]['email'])
    for msg_id in search_data[0].split():
        # msg_id_str = msg_id.decode('utf-8')
        # print("Fetching message {} of {}".format(msg_id_str, nmessages))
        status, msg_data = imap4.fetch(msg_id, '(RFC822)')
        # print(msg_data)
        # print(msg_data[0])
        # print(msg_data[0][1])
        # print(type(msg_data[0][1]))
        # msg_raw = msg_data
        # print(msg_raw)
        msg = email.message_from_bytes(msg_data[0][1], _class=email.message.EmailMessage)
        # print(msg)
        # print(msg['Subject'], msg['X-Mailing-List'], msg['Date'], email.utils.parsedate_tz(msg['Date']))
        payload = msg.get_payload()
        print(payload)
        for f in msg.get_payload():
            if f.get_filename():
                filename = email.header.decode_header(f.get_filename())
                if filename[0][1]:
                    filename = './mail/' + js_d[0]['client'] + '/' + time.strftime("%b_%d") + '/' + \
                               filename[0][0].decode(filename[0][1])
                    dir_name = os.path.dirname(filename)
                    subprocess.call('mkdir -p {}'.format(dir_name), shell=True)
                    print(filename + 'decode')
                    with open(filename, 'wb') as new_file:
                        new_file.write(f.get_payload(decode=True))
                    text.write(time.strftime("%b_%d_%Y_%H_%M_") + '  ' + filename + '\n')
                    re_name(js_d, filename)
                else:
                    filename = filename[0][0]
                    print(filename + 'not_decode')
                    filename = './mail/' + js_d[0]['client'] + '/' + time.strftime("%b_%d") + '/' + \
                               filename
                    dir_name = os.path.dirname(filename)
                    subprocess.call('mkdir -p {}'.format(dir_name), shell=True)
                    print(filename)
                    with open(filename, 'wb') as new_file:
                        new_file.write(f.get_payload(decode=True))
                    text.write(time.strftime("%b_%d_%Y_%H_%M_") + '  ' + filename + '\n')
                    re_name(js_d, filename)

        result = imap4.copy(msg_id, 'INBOX.old_1')
        imap4.store(msg_id, '+FLAGS', '\\Deleted')
        print(result)
    imap4.expunge()
    imap4.close()
    imap4.logout()




def try_p(pharm):
        try:
            main_loop(pharm)
        except Exception as e:
            print("ERROR:" + str(e))
            text.write('\n' + time.strftime("%b_%d_%Y_%H_%M_") + '  ' + "ERROR:" + str(e) + '\n' + '\n')
        # print("Sleeping {} seconds...".format(pause_time))
        # time.sleep(pause_time)


def pharmacy_name(pharmacy):
    for p in json_f(pharmacy):
        json_p = json_f(pharmacy)[p]
        text.write('\n' + time.strftime("%b_%d_%Y_%H_%M_") + p + '\n')
        print(json_p)
        try_p(json_p)
        print('finish')


pharmacy_name(Pharmacy)

text.close()
